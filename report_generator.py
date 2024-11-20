import pandas as pd
import openpyxl


# Функция для автоподбора ширины столбцов
def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Немного увеличиваем ширину
        ws.column_dimensions[column].width = adjusted_width

# Функция для сохранения в Excel с добавлением формул для сумм
def save_to_excel_with_totals(df, excel_file, total_column, total_label):
    df.to_excel(excel_file, index=False)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Подбор ширины столбцов
    adjust_column_widths(ws)

    # Получаем номер строки для вставки итогов
    total_row = len(df) + 2
    ws[f"A{total_row}"] = total_label
    ws[f"E{total_row}"] = f"=SUM(E2:E{total_row - 1})"  # Сумма по столбцу "E"

    wb.save(excel_file)


# Отчет по продажам
def generate_sales_report_excel(data):
    if isinstance(data, list):
        data = pd.DataFrame(data)

    columns_mapping = {
        "date_from": "Дата начала",
        "date_to": "Дата конца",
        "create_dt": "Дата создания",
        "subject_name": "Предмет",
        "sa_name": "Артикул продавца",
        "barcode": "Баркод",
        "shk_id": "Штрих-код",
        "ppvz_for_pay": "К перечислению"
    }

    missing_columns = [col for col in columns_mapping.keys() if col not in data.columns]
    if missing_columns:
        raise ValueError(f"Отсутствуют необходимые столбцы: {', '.join(missing_columns)}")

    sales_data = data[data["supplier_oper_name"] == "Продажа"]
    logistics_data = data[data["supplier_oper_name"] == "Логистика"]

    df_sales = sales_data[columns_mapping.keys()].rename(columns=columns_mapping)

    # Сохранение отчета с добавлением итога
    excel_file = "Отчет по продажам.xlsx"
    logistics_total = logistics_data["delivery_rub"].sum()
    save_to_excel_with_totals(df_sales, excel_file, "ppvz_for_pay", "Итого")

    # Открываем и добавляем итоговую сумму по логистике
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    total_row = len(df_sales) + 2
    ws[f"H{total_row}"] = f"=SUM(H2:H{total_row - 1}) - {logistics_total}"

    wb.save(excel_file)
    return excel_file


# Отчет об остатках на складе
def generate_stock_report_excel(data):
    data = pd.DataFrame(data)
    columns_mapping = {
        "supplierArticle": "Артикул",
        "subject": "Предмет",
        "quantity": "На складе",
        "inWayToClient": "В пути к клиенту",
        "inWayFromClient": "От клиента",
        "techSize": "Размер",
        "Price": "Цена"
    }
    df_stock = data[list(columns_mapping.keys())].rename(columns=columns_mapping)

    excel_file = "ОстаткиНаСкладе.xlsx"
    df_stock.to_excel(excel_file, index=False)
    return excel_file


# Отчет о заказах
def generate_orders_report_excel(data):
    data = pd.DataFrame(data)
    columns_mapping = {
        "date": "Дата заказа",
        "lastChangeDate": "Изменение даты",

        "supplierArticle": "Артикул",
        "barcode": "Баркод",
        "finishedPrice": "Цена со скидкой",
        "isCancel": "Отменён",
        "orderType": "Тип заказа",
        "sticker": "ШК",
    }

    df_orders = data[list(columns_mapping.keys())].rename(columns=columns_mapping)
    df_orders["Отменён"] = df_orders["Отменён"].apply(lambda x: "Да" if x else "Нет")

    excel_file = "Отчет о заказах.xlsx"
    df_orders.to_excel(excel_file, index=False)
    return excel_file


# Отчет о продажах и возвратах
def generate_sales_and_returns_report_excel(data):
    data = pd.DataFrame(data)
    columns_mapping = {
        "subject": "Предмет",
        "supplierArticle": "Артикул продавца",
        "barcode": "Баркод",
        "totalPrice": "Цена без скидок",
        "discountPercent": "Скидка продавца",
        "forPay": "К перечислению",
        "priceWithDisc": "Цена со скидкой",
        "saleID": "ID продажи",
        "date": "Дата продажи"
    }
    df_sales = data[list(columns_mapping.keys())].rename(columns=columns_mapping)

    excel_file = "продажи_возвраты.xlsx"
    df_sales.to_excel(excel_file, index=False)
    return excel_file
