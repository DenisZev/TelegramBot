import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import logging
from wildberries_api import get_stock_data, get_sales_data, get_sales_report, get_orders_data
import numpy as np

async def create_final_report(date_from, date_to):
    try:
        # Преобразуем даты в формат datetime без часового пояса
        date_from = pd.to_datetime(date_from, errors='coerce')
        date_to = pd.to_datetime(date_to, errors='coerce')

        # Проверка, что дата преобразована корректно
        if pd.isnull(date_from) or pd.isnull(date_to):
            raise ValueError("Неверный формат дат. Пожалуйста, укажите правильные значения для 'date_from' и 'date_to'.")

        # Преобразуем даты в строковый формат 'YYYY-MM-DD' для использования в API
        date_from_str = date_from.strftime('%Y-%m-%d')
        date_to_str = date_to.strftime('%Y-%m-%d')

        # Получаем данные
        stock_data = await get_stock_data(date_from_str)
        sales_report = await get_sales_report(date_from_str, date_to_str)
        orders_data = await get_orders_data(date_from_str, date_to_str)
        sales_data = await get_sales_data(date_from_str)

        # Проверка наличия данных
        if not stock_data or not sales_report or not orders_data or not sales_data:
            raise ValueError("Одни или несколько из запрашиваемых данных отсутствуют. Проверьте API.")

        # Преобразуем данные в DataFrame
        df_sales = pd.DataFrame(sales_data)
        df_orders = pd.DataFrame(orders_data)
        df_stock = pd.DataFrame(stock_data)
        df_sales_report = pd.DataFrame(sales_report)

        # Переименование столбцов для согласованности
        rename_sales = {'forPay': 'ppvz_for_pay', 'lastChangeDate': 'sales_lastChangeDate'}
        rename_orders = {'lastChangeDate': 'order_lastChangeDate', 'finishedPrice': 'order_price'}
        rename_stock = {'lastChangeDate': 'stock_lastChangeDate'}
        rename_sales_report = {'sa_name': 'supplierArticle'}

        df_sales.rename(columns=rename_sales, inplace=True)
        df_orders.rename(columns=rename_orders, inplace=True)
        df_stock.rename(columns=rename_stock, inplace=True)
        df_sales_report.rename(columns=rename_sales_report, inplace=True)

        # Преобразуем все даты в формате UTC в локальное время или в без часового пояса
        df_sales['sales_lastChangeDate'] = pd.to_datetime(df_sales['sales_lastChangeDate'], errors='coerce').dt.tz_localize(None)
        df_orders['order_lastChangeDate'] = pd.to_datetime(df_orders['order_lastChangeDate'], errors='coerce').dt.tz_localize(None)
        df_sales_report['sale_dt'] = pd.to_datetime(df_sales_report['sale_dt'], errors='coerce').dt.tz_localize(None)

        # Фильтрация данных по дате
        df_sales = df_sales[(df_sales['sales_lastChangeDate'] >= date_from) & (df_sales['sales_lastChangeDate'] <= date_to)]
        df_orders = df_orders[(df_orders['order_lastChangeDate'] >= date_from) & (df_orders['order_lastChangeDate'] <= date_to)]
        df_sales_report = df_sales_report[(df_sales_report['sale_dt'] >= date_from) & (df_sales_report['sale_dt'] <= date_to)]

        # Логистика данные и расчет итога
        logistics_data = df_sales_report[df_sales_report["supplier_oper_name"] == "Логистика"]
        logistics_total = logistics_data["delivery_rub"].sum()

        # Объединение данных по артикулу и srid (без логистики)
        combined_data = pd.merge(
            df_sales[['supplierArticle', 'totalPrice', 'ppvz_for_pay', 'srid', 'sales_lastChangeDate']],
            df_orders[['supplierArticle', 'order_price', 'isCancel', 'srid', 'order_lastChangeDate']],
            on=['supplierArticle', 'srid'],
            how='outer'
        )

        # Логирование информации об объединении данных
        logging.info(f"Объединено {len(combined_data)} записей из продаж и заказов.")

        # Суммирование по 'srid' и 'supplierArticle'
        combined_data['ppvz_for_pay'] = combined_data.groupby('srid')['ppvz_for_pay'].transform('sum')

        # Заполнение пропусков и преобразование данных
        combined_data.fillna({
            'ppvz_for_pay': 0,
            'totalPrice': 0,
            'sales_lastChangeDate': pd.NaT,
            'order_price': 0,
            'isCancel': False,
            'order_lastChangeDate': pd.NaT
        }, inplace=True)

        # Логируем количество заполненных пропусков
        logging.info(f"Заполнено {combined_data.isnull().sum().sum()} пропусков в данных.")

        # Обработка строк без дат, заполняем их значениями
        combined_data['sales_lastChangeDate'].fillna(date_from, inplace=True)
        combined_data['order_lastChangeDate'].fillna(date_from, inplace=True)

        # Определение статуса с помощью np.select для оптимизации
        combined_data["Статус"] = np.select(
            [
                combined_data["ppvz_for_pay"].notnull() & ~combined_data["isCancel"],
                ~combined_data["isCancel"]
            ],
            ["Продан", "В пути"],
            default="Возвращен"
        )

        # Сортировка по дате изменения продаж
        combined_data['sales_lastChangeDate'] = pd.to_datetime(combined_data['sales_lastChangeDate'], errors='coerce')
        combined_data = combined_data.sort_values(by='sales_lastChangeDate')

        # Создание Excel-файла
        excel_file = "Итоговый отчет.xlsx"
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Запись данных на отдельные листы
            combined_data[['sales_lastChangeDate', 'supplierArticle', 'ppvz_for_pay', 'order_price', 'Статус', 'srid']].to_excel(writer, sheet_name="Продажи и Заказы", index=False)
            df_stock[['supplierArticle', 'quantity', 'quantityFull']].to_excel(writer, sheet_name="Остатки на складе", index=False)
            combined_data[combined_data["Статус"] == "В пути"].to_excel(writer, sheet_name="В пути", index=False)
            combined_data[combined_data["Статус"] == "Возвращен"].to_excel(writer, sheet_name="Возвращен", index=False)
            combined_data[combined_data["Статус"] == "Продан"].to_excel(writer, sheet_name="Продано", index=False)

            # Проверим, существует ли поле "subject" в данных по логистике перед записью
            if 'subject' in logistics_data.columns:
                logistics_data[['supplierArticle', 'delivery_rub', 'srid', 'subject']].to_excel(writer, sheet_name="Логистика", index=False)
            else:
                logistics_data[['supplierArticle', 'delivery_rub', 'srid']].to_excel(writer, sheet_name="Логистика", index=False)

        # Условное форматирование
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Продажи и Заказы"]

        # Определение индекса столбца "Статус" для применения цветового форматирования
        status_col_index = [cell.value for cell in ws[1]].index("Статус") + 1

        # Цветовые заливки для каждого статуса
        fills = {
            "Продан": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
            "В пути": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
            "Возвращен": PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
        }

        # Применяем форматирование на основе статуса
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=status_col_index, max_col=status_col_index):
            for cell in row:
                cell.fill = fills.get(cell.value, PatternFill())

        # Сохраняем файл с форматированием
        wb.save(excel_file)

        # Возвращаем путь к файлу и общую сумму логистики
        return excel_file, logistics_total

    except Exception as e:
        logging.error(f"Ошибка при создании итогового отчета: {e}")
        raise
