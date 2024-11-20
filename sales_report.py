import pandas as pd
import openpyxl
from openpyxl.styles import Font

def g_sales_report_excel(data):
    # Словарь для выбора и переименования нужных столбцов
    columns_mapping = {
        "subject": "Предмет",
        "supplierArticle": "Артикул продавца",
        "barcode": "Баркод",
        "totalPrice": "Цена без скидок",
        "discountPercent": "Скидка продавца",
        "forPay": "К перечислению",
        "priceWithDisc": "Цена со скидкой",
        "saleID": "ID продажи",
        "date": "Дата продажи"
    }

    # Создаем DataFrame только для продаж, исключая возвраты
    df_sales = pd.DataFrame(data)
    df_sales = df_sales[df_sales["saleID"].str.startswith("S")]  # Оставляем только продажи
    df_sales = df_sales[columns_mapping.keys()].rename(columns=columns_mapping)

    # Сохраняем отчет в Excel
    excel_file = "sales_report.xlsx"
    df_sales.to_excel(excel_file, index=False, startrow=0, sheet_name="Sales Report")

    # Открываем Excel-файл для добавления итоговой суммы
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Добавляем строку с итоговой суммой "К перечислению"
    total_row = len(df_sales) + 2
    ws[f"A{total_row}"] = "Итого"
    ws[f"F{total_row}"] = f"=SUM(F2:F{total_row - 1})"  # Итоговая сумма "К перечислению"

    # Применяем жирный шрифт к итоговой строке
    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"F{total_row}"].font = Font(bold=True)

    wb.save(excel_file)
    return excel_file
